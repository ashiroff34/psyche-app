#!/usr/bin/env python3
"""
excel_to_ts.py — Reads Thyself_Questions_Scoring.xlsx and regenerates
all TypeScript data files for the Thyself app.

Usage:
    python3 scripts/excel_to_ts.py
    python3 scripts/excel_to_ts.py --excel /path/to/custom.xlsx
    python3 scripts/excel_to_ts.py --dry-run   # print TS to stdout, don't write files

Edit questions directly in the Excel file, then run this script to push changes into the app.
"""

import argparse
import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

# ── Paths ────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
REPO_ROOT   = SCRIPT_DIR.parent
DATA_DIR    = REPO_ROOT / "src" / "data"
ASSESS_DIR  = DATA_DIR / "assessments"
EXCEL_PATH  = Path.home() / "Documents" / "Thyself_Questions_Scoring.xlsx"

OUTPUT_FILES = {
    "type_quizzes":          DATA_DIR / "type-quizzes.ts",
    "essential_enneagram":   ASSESS_DIR / "essential-enneagram.ts",
    "ieq9":                  ASSESS_DIR / "ieq9-style.ts",
    "cognitive":             ASSESS_DIR / "cognitive-type.ts",
    "big_five":              ASSESS_DIR / "big-five.ts",
    "personality_path":      ASSESS_DIR / "personality-path.ts",
    "michael_caloz":         ASSESS_DIR / "michael-caloz.ts",
    "mistype":               ASSESS_DIR / "mistype-investigator.ts",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def ts_str(s):
    """Escape a string for TypeScript."""
    if s is None:
        return '""'
    s = str(s).replace("\\", "\\\\").replace("`", "\\`").replace("${", "\\${")
    # Use single-quoted TS strings; escape single quotes
    s = s.replace("'", "\\'")
    return f"'{s}'"

def ts_num(v, default=0):
    try:
        return int(v) if v is not None else default
    except (ValueError, TypeError):
        return default

def load_sheet(wb, name):
    """Return all non-empty rows starting from row 3 (skip header + blank)."""
    ws = wb[name]
    rows = []
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))
    return rows

# ── Sheet parsers → TypeScript generators ─────────────────────────────────────

def gen_type_quizzes(wb):
    rows = load_sheet(wb, "Daily_Path_Quizzes")
    lines = [
        "export interface TypeQuizQuestion {",
        "  id: number;",
        "  type: number;",
        "  difficulty: 1 | 2 | 3 | 4 | 5;",
        "  track1Day: number;",
        "  track2Day: number;",
        "  category: string;",
        "  question: string;",
        '  options: { letter: "A" | "B" | "C" | "D"; text: string }[];',
        '  answer: "A" | "B" | "C" | "D";',
        "  explanation: string;",
        "}",
        "",
        "export const typeQuizQuestions: TypeQuizQuestion[] = [",
    ]
    for r in rows:
        diff = ts_num(r.get("Difficulty (1-5)"), 1)
        lines.append("  {")
        lines.append(f"    id: {ts_num(r['ID'])},")
        lines.append(f"    type: {ts_num(r['Type'])},")
        lines.append(f"    difficulty: {diff} as 1 | 2 | 3 | 4 | 5,")
        lines.append(f"    track1Day: {ts_num(r.get('Track1_Day'))},")
        lines.append(f"    track2Day: {ts_num(r.get('Track2_Day'))},")
        lines.append(f"    category: {ts_str(r.get('Category'))},")
        lines.append(f"    question: {ts_str(r.get('Question'))},")
        lines.append("    options: [")
        for letter in ["A", "B", "C", "D"]:
            text = r.get(f"Option_{letter}", "")
            lines.append(f'      {{ letter: "{letter}", text: {ts_str(text)} }},')
        lines.append("    ],")
        lines.append(f"    answer: \"{r.get('Answer', 'A')}\",")
        lines.append(f"    explanation: {ts_str(r.get('Explanation'))},")
        lines.append("  },")
    lines.append("];")
    return "\n".join(lines)


def gen_essential_enneagram(wb):
    ws = wb["Essential_Enneagram"]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Find section boundaries
    para_rows, nq_rows = [], []
    section = None
    headers = None
    for row in all_rows:
        if row[0] and "SECTION A" in str(row[0]):
            section = "para"
            headers = None
        elif row[0] and "SECTION B" in str(row[0]):
            section = "nq"
            headers = None
        elif row[0] in ("Type_Number", "Types_Compared"):
            headers = row
        elif headers and any(v is not None for v in row):
            d = dict(zip(headers, row))
            if section == "para":
                para_rows.append(d)
            elif section == "nq":
                nq_rows.append(d)

    lines = [
        "export interface TypeParagraph {",
        "  typeNumber: number;",
        "  paragraph: string;",
        "}",
        "",
        "export interface NarrowingQuestion {",
        "  types: number[];",
        "  question: string;",
        "  options: { text: string; scores: Record<number, number> }[];",
        "}",
        "",
        "export const essentialParagraphs: TypeParagraph[] = [",
    ]
    for r in para_rows:
        lines.append("  {")
        lines.append(f"    typeNumber: {ts_num(r.get('Type_Number'))},")
        lines.append(f"    paragraph: {ts_str(r.get('Paragraph_Text'))},")
        lines.append("  },")
    lines.append("];")
    lines.append("")
    lines.append("export const narrowingQuestions: NarrowingQuestion[] = [")

    # Group narrowing questions by question text (each row = one option)
    from collections import OrderedDict
    q_map = OrderedDict()
    for r in nq_rows:
        q_text = str(r.get("Question", "") or "")
        types_str = str(r.get("Types_Compared", "") or "")
        types = [int(x.strip()) for x in types_str.replace("vs", ",").split(",") if x.strip().isdigit()]
        if q_text not in q_map:
            q_map[q_text] = {"types": types, "options": []}
        opt_text = r.get("Option_Text") or r.get("Option_1_Text") or ""
        # collect all score columns
        scores = {}
        for k, v in r.items():
            if k and "Score_T" in str(k) and v is not None:
                t = k.replace("Score_T", "").strip()
                if t.isdigit():
                    scores[int(t)] = int(v)
        if opt_text:
            q_map[q_text]["options"].append({"text": str(opt_text), "scores": scores})

    for q_text, q_data in q_map.items():
        lines.append("  {")
        types_ts = ", ".join(str(t) for t in q_data["types"])
        lines.append(f"    types: [{types_ts}],")
        lines.append(f"    question: {ts_str(q_text)},")
        lines.append("    options: [")
        for opt in q_data["options"]:
            scores_ts = ", ".join(f"{k}: {v}" for k, v in opt["scores"].items())
            lines.append(f"      {{ text: {ts_str(opt['text'])}, scores: {{ {scores_ts} }} }},")
        lines.append("    ],")
        lines.append("  },")
    lines.append("];")
    return "\n".join(lines)


def gen_ieq9(wb):
    rows = load_sheet(wb, "IEQ9_Likert")
    phase_map = {
        "Phase 1 – Core Type": "core",
        "Phase 2 – Wing Refinement": "wing",
        "Phase 3 – Instinctual Variants": "instinct",
        "Phase 4 – Stress/Growth": "stress-growth",
    }

    lines = [
        "export interface LikertStatement {",
        "  id: number;",
        "  text: string;",
        "  scores: Record<string, number>;",
        "  category: 'core' | 'wing' | 'instinct' | 'stress-growth';",
        "}",
        "",
    ]

    # Group by category for separate exports
    by_cat = {"core": [], "wing": [], "instinct": [], "stress-growth": []}
    for r in rows:
        phase_str = str(r.get("Phase") or "")
        cat_raw = str(r.get("Category") or "").strip()
        # Use Category column if it's one of the known values, else derive from Phase
        if cat_raw in by_cat:
            cat = cat_raw
        else:
            cat = next((v for k, v in phase_map.items() if k in phase_str), "core")
        by_cat[cat].append(r)

    export_names = {
        "core": "coreStatements",
        "wing": "wingStatements",
        "instinct": "instinctStatements",
        "stress-growth": "stressGrowthStatements",
    }

    for cat, export_name in export_names.items():
        lines.append(f"export const {export_name}: LikertStatement[] = [")
        for r in by_cat[cat]:
            scores = {}
            for t in range(1, 10):
                v = r.get(f"Type_{t}_Weight")
                if v is not None and str(v).strip() not in ("", "None", "0"):
                    scores[str(t)] = int(float(str(v)))
            scores_ts = ", ".join(f"'{k}': {v}" for k, v in scores.items())
            lines.append("  {")
            lines.append(f"    id: {ts_num(r['ID'])},")
            lines.append(f"    text: {ts_str(r.get('Statement'))},")
            lines.append(f"    scores: {{ {scores_ts} }},")
            lines.append(f"    category: '{cat}',")
            lines.append("  },")
        lines.append("];")
        lines.append("")

    return "\n".join(lines)


def gen_cognitive(wb):
    rows = load_sheet(wb, "Cognitive_Type")
    func_keys = ["Se", "Si", "Ne", "Ni", "Te", "Ti", "Fe", "Fi"]

    lines = [
        "export interface CognitiveItem {",
        "  id: number;",
        "  text: string;",
        "  scores: Record<string, number>;",
        "  category: 'perceiving' | 'judging';",
        "}",
        "",
        "export const cognitiveTypeItems: CognitiveItem[] = [",
    ]
    for r in rows:
        scores = {}
        for f in func_keys:
            v = r.get(f"{f}_Weight")
            if v is not None and str(v).strip() not in ("", "None", "0"):
                scores[f] = int(float(str(v)))
        scores_ts = ", ".join(f"{k}: {v}" for k, v in scores.items())
        cat = str(r.get("Category") or "perceiving").strip()
        lines.append("  {")
        lines.append(f"    id: {ts_num(r['ID'])},")
        lines.append(f"    text: {ts_str(r.get('Statement'))},")
        lines.append(f"    scores: {{ {scores_ts} }},")
        lines.append(f"    category: '{cat}',")
        lines.append("  },")
    lines.append("];")
    lines += [
        "",
        "export const typeStacks: Record<string, string[]> = {",
        "  INTJ: ['Ni', 'Te', 'Fi', 'Se'],",
        "  INTP: ['Ti', 'Ne', 'Si', 'Fe'],",
        "  ENTJ: ['Te', 'Ni', 'Se', 'Fi'],",
        "  ENTP: ['Ne', 'Ti', 'Fe', 'Si'],",
        "  INFJ: ['Ni', 'Fe', 'Ti', 'Se'],",
        "  INFP: ['Fi', 'Ne', 'Si', 'Te'],",
        "  ENFJ: ['Fe', 'Ni', 'Se', 'Ti'],",
        "  ENFP: ['Ne', 'Fi', 'Te', 'Si'],",
        "  ISTJ: ['Si', 'Te', 'Fi', 'Ne'],",
        "  ISFJ: ['Si', 'Fe', 'Ti', 'Ne'],",
        "  ESTJ: ['Te', 'Si', 'Ne', 'Fi'],",
        "  ESFJ: ['Fe', 'Si', 'Ne', 'Ti'],",
        "  ISTP: ['Ti', 'Se', 'Ni', 'Fe'],",
        "  ISFP: ['Fi', 'Se', 'Ni', 'Te'],",
        "  ESTP: ['Se', 'Ti', 'Fe', 'Ni'],",
        "  ESFP: ['Se', 'Fi', 'Te', 'Ni'],",
        "};",
    ]
    return "\n".join(lines)


def gen_big_five(wb):
    rows = load_sheet(wb, "Big_Five")
    factor_map = {
        "O – Openness to Experience": "O",
        "C – Conscientiousness": "C",
        "E – Extraversion": "E",
        "A – Agreeableness": "A",
        "N – Neuroticism": "N",
    }

    lines = [
        "export interface BigFiveItem {",
        "  id: number;",
        "  text: string;",
        "  factor: 'O' | 'C' | 'E' | 'A' | 'N';",
        "  facet: string;",
        "  reversed: boolean;",
        "}",
        "",
        "export const bigFiveItems: BigFiveItem[] = [",
    ]
    for r in rows:
        factor_raw = str(r.get("Factor") or "O")
        factor = factor_map.get(factor_raw, factor_raw[0] if factor_raw else "O")
        rev_raw = str(r.get("Reversed") or "No").strip().lower()
        reversed_val = "true" if rev_raw in ("yes", "true", "1") else "false"
        lines.append(
            f"  {{ id: {ts_num(r['ID'])}, text: {ts_str(r.get('Statement'))}, "
            f"factor: '{factor}', facet: {ts_str(r.get('Facet'))}, reversed: {reversed_val} }},"
        )
    lines.append("];")
    lines += [
        "",
        'export const factorDescriptions: Record<string, { name: string; high: string; low: string; description: string }> = {',
        '  O: {',
        '    name: "Openness to Experience",',
        '    high: "Imaginative, curious, and open to new ideas and experiences. Appreciates art, emotion, adventure, and unconventional perspectives.",',
        '    low: "Practical, conventional, and grounded. Prefers familiar routines and straightforward, concrete thinking.",',
        '    description: "Openness to Experience reflects the degree to which a person is intellectually curious, creative, and receptive to novel ideas, feelings, and experiences. High scorers tend to be imaginative and independent-minded; low scorers tend to be pragmatic and traditional.",',
        '  },',
        '  C: {',
        '    name: "Conscientiousness",',
        '    high: "Disciplined, organized, and dependable. Sets clear goals and follows through with determination and careful planning.",',
        '    low: "Flexible, spontaneous, and easygoing. May be less structured and more prone to acting on impulse.",',
        '    description: "Conscientiousness describes the tendency to be organized, responsible, and goal-directed. High scorers are reliable and hardworking; low scorers tend to be more laid-back and may prefer flexibility over structure.",',
        '  },',
        '  E: {',
        '    name: "Extraversion",',
        '    high: "Outgoing, energetic, and sociable. Enjoys being around people, seeks excitement, and expresses positive emotions easily.",',
        '    low: "Reserved, quiet, and independent. Prefers solitary activities and less stimulating environments.",',
        '    description: "Extraversion reflects the extent to which a person is energized by social interaction and external stimulation. High scorers are talkative, assertive, and enthusiastic; low scorers are more introspective and prefer solitude.",',
        '  },',
        '  A: {',
        '    name: "Agreeableness",',
        '    high: "Compassionate, cooperative, and trusting. Values getting along with others and tends to be considerate and helpful.",',
        '    low: "Competitive, skeptical, and challenging. Prioritizes personal interests and is more willing to engage in conflict.",',
        '    description: "Agreeableness reflects the tendency toward compassion, cooperation, and social harmony. High scorers are trusting and altruistic; low scorers are more analytical, competitive, and willing to challenge others.",',
        '  },',
        '  N: {',
        '    name: "Neuroticism",',
        '    high: "Sensitive to stress and prone to experiencing negative emotions such as anxiety, sadness, and irritability.",',
        '    low: "Emotionally stable, calm, and resilient. Handles stress well and rarely feels upset or overwhelmed.",',
        '    description: "Neuroticism describes the tendency to experience negative emotions and psychological distress. High scorers are more reactive to stress and experience more frequent mood fluctuations; low scorers are emotionally stable and even-tempered.",',
        '  },',
        '};',
    ]
    return "\n".join(lines)


def gen_personality_path(wb):
    rows = load_sheet(wb, "Personality_Path")

    lines = [
        "export interface AdaptiveRound {",
        "  id: string;",
        "  title: string;",
        "  description: string;",
        "  questions: {",
        "    text: string;",
        "    options: { text: string; scores: Record<string, number> }[];",
        "  }[];",
        "}",
        "",
        "export const adaptiveRounds: AdaptiveRound[] = [",
    ]

    from collections import OrderedDict
    round_map = OrderedDict()
    for r in rows:
        rid = str(r.get("Round_ID") or "")
        if not rid:
            continue
        if rid not in round_map:
            round_map[rid] = {
                "title": r.get("Round_Title", ""),
                "description": r.get("Round_Description", ""),
                "questions": OrderedDict(),
            }
        q_text = str(r.get("Question") or "")
        if q_text not in round_map[rid]["questions"]:
            round_map[rid]["questions"][q_text] = []
        # Parse options 1-3
        for i in range(1, 4):
            opt_text = r.get(f"Option_{i}_Text")
            opt_scores_raw = r.get(f"Option_{i}_Scores")
            if opt_text:
                scores = {}
                if opt_scores_raw:
                    # Format: "8:2, 9:2, 1:2" or JSON
                    raw = str(opt_scores_raw).strip()
                    if raw.startswith("{"):
                        try:
                            scores = json.loads(raw)
                        except Exception:
                            pass
                    else:
                        for part in raw.split(","):
                            if ":" in part:
                                k, v = part.split(":", 1)
                                try:
                                    scores[k.strip().strip('"')] = int(v.strip())
                                except ValueError:
                                    pass
                round_map[rid]["questions"][q_text].append({"text": str(opt_text), "scores": scores})

    for rid, rdata in round_map.items():
        lines.append("  {")
        lines.append(f"    id: {ts_str(rid)},")
        lines.append(f"    title: {ts_str(rdata['title'])},")
        lines.append(f"    description: {ts_str(rdata['description'])},")
        lines.append("    questions: [")
        for q_text, opts in rdata["questions"].items():
            lines.append("      {")
            lines.append(f"        text: {ts_str(q_text)},")
            lines.append("        options: [")
            for opt in opts:
                scores_ts = ", ".join(f'"{k}": {v}' for k, v in opt["scores"].items())
                lines.append(f"          {{ text: {ts_str(opt['text'])}, scores: {{ {scores_ts} }} }},")
            lines.append("        ],")
            lines.append("      },")
        lines.append("    ],")
        lines.append("  },")
    lines.append("];")
    return "\n".join(lines)


def gen_michael_caloz(wb):
    rows = load_sheet(wb, "Michael_Caloz")
    lines = [
        "export interface CalozSection {",
        "  id: number;",
        "  theme: string;",
        "  description: string;",
        "  statements: { type: number; text: string }[];",
        "}",
        "",
        "export const calozSections: CalozSection[] = [",
    ]

    from collections import OrderedDict
    sec_map = OrderedDict()
    for r in rows:
        sid = ts_num(r.get("Section_ID"))
        if sid not in sec_map:
            sec_map[sid] = {
                "theme": r.get("Theme", ""),
                "description": r.get("Section_Description", ""),
                "statements": [],
            }
        sec_map[sid]["statements"].append({
            "type": ts_num(r.get("Type_Number")),
            "text": str(r.get("Statement") or ""),
        })

    for sid, sdata in sec_map.items():
        lines.append("  {")
        lines.append(f"    id: {sid},")
        lines.append(f"    theme: {ts_str(sdata['theme'])},")
        lines.append(f"    description: {ts_str(sdata['description'])},")
        lines.append("    statements: [")
        for stmt in sdata["statements"]:
            lines.append(f"      {{ type: {stmt['type']}, text: {ts_str(stmt['text'])} }},")
        lines.append("    ],")
        lines.append("  },")
    lines.append("];")
    return "\n".join(lines)


def gen_mistype(wb):
    rows = load_sheet(wb, "Mistype_Investigator")
    lines = [
        "export interface MistypePair {",
        "  typeA: number;",
        "  typeB: number;",
        "  commonConfusion: string;",
        "  questions: {",
        "    text: string;",
        "    optionA: { text: string; leansToward: number };",
        "    optionB: { text: string; leansToward: number };",
        "  }[];",
        "}",
        "",
        "export const mistypePairs: MistypePair[] = [",
    ]

    from collections import OrderedDict
    pair_map = OrderedDict()
    for r in rows:
        key = (ts_num(r.get("Type_A")), ts_num(r.get("Type_B")))
        if key not in pair_map:
            pair_map[key] = {
                "confusion": str(r.get("Common_Confusion") or ""),
                "questions": [],
            }
        q = {
            "text": str(r.get("Question") or ""),
            "optA": str(r.get("Option_A_Text") or ""),
            "leansA": ts_num(r.get("Option_A_Leans_Toward")),
            "optB": str(r.get("Option_B_Text") or ""),
            "leansB": ts_num(r.get("Option_B_Leans_Toward")),
        }
        if q["text"]:
            pair_map[key]["questions"].append(q)

    for (tA, tB), pdata in pair_map.items():
        lines.append("  {")
        lines.append(f"    typeA: {tA},")
        lines.append(f"    typeB: {tB},")
        lines.append(f"    commonConfusion: {ts_str(pdata['confusion'])},")
        lines.append("    questions: [")
        for q in pdata["questions"]:
            lines.append("      {")
            lines.append(f"        text: {ts_str(q['text'])},")
            lines.append(f"        optionA: {{ text: {ts_str(q['optA'])}, leansToward: {q['leansA']} }},")
            lines.append(f"        optionB: {{ text: {ts_str(q['optB'])}, leansToward: {q['leansB']} }},")
            lines.append("      },")
        lines.append("    ],")
        lines.append("  },")
    lines.append("];")
    return "\n".join(lines)


# ── Main ──────────────────────────────────────────────────────────────────────

GENERATORS = {
    "type_quizzes":        ("Daily_Path_Quizzes",    gen_type_quizzes),
    "essential_enneagram": ("Essential_Enneagram",   gen_essential_enneagram),
    "ieq9":                ("IEQ9_Likert",            gen_ieq9),
    "cognitive":           ("Cognitive_Type",         gen_cognitive),
    "big_five":            ("Big_Five",               gen_big_five),
    "personality_path":    ("Personality_Path",       gen_personality_path),
    "michael_caloz":       ("Michael_Caloz",          gen_michael_caloz),
    "mistype":             ("Mistype_Investigator",   gen_mistype),
}

def main():
    parser = argparse.ArgumentParser(description="Regenerate TS data files from Excel")
    parser.add_argument("--excel", default=str(EXCEL_PATH), help="Path to Excel file")
    parser.add_argument("--dry-run", action="store_true", help="Print output, don't write files")
    parser.add_argument("--only", help="Run only one generator (e.g. type_quizzes)")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"ERROR: Excel file not found: {excel_path}")
        sys.exit(1)

    print(f"Loading {excel_path.name}...")
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

    for key, (sheet_name, gen_fn) in GENERATORS.items():
        if args.only and args.only != key:
            continue
        if sheet_name not in wb.sheetnames:
            print(f"  SKIP {key} — sheet '{sheet_name}' not found")
            continue
        try:
            ts_content = gen_fn(wb)
            out_path = OUTPUT_FILES[key]
            if args.dry_run:
                print(f"\n{'='*60}")
                print(f"  {key} → {out_path.name}")
                print('='*60)
                print(ts_content[:500] + "...\n")
            else:
                out_path.parent.mkdir(parents=True, exist_ok=True)
                out_path.write_text(ts_content, encoding="utf-8")
                lines = ts_content.count("\n")
                print(f"  ✓ {key} → {out_path.relative_to(REPO_ROOT)}  ({lines} lines)")
        except Exception as e:
            import traceback
            print(f"  ✗ {key} — ERROR: {e}")
            traceback.print_exc()

    print("\nDone. Run `npm run build` in the app to verify TypeScript compiles cleanly.")

if __name__ == "__main__":
    main()
